"""Normalize raw vehicle rows into the recommendation schema."""

from __future__ import annotations

from collections import Counter, defaultdict
import json
from pathlib import Path
from statistics import median
import re
from typing import Any, Dict, Iterable, List, Mapping, Optional, Sequence, Tuple

RETAIN_COLUMNS = [
    "Make and model",
    "Vehicle year",
    "sub model",
    "Body style",
    "Fuel type",
    "transmission",
    "engine size",
    "doors",
    "seats",
    "safety stars",
    "safety rating",
    "fuel consumption (litres per 100km)",
    "fuel stars",
]

BRAND_REGIONS = {
    "audi": "European",
    "bmw": "European",
    "citroen": "European",
    "fiat": "European",
    "jaguar": "European",
    "land rover": "European",
    "mercedes-benz": "European",
    "mercedes": "European",
    "mini": "European",
    "peugeot": "European",
    "renault": "European",
    "skoda": "European",
    "volkswagen": "European",
    "vw": "European",
    "volvo": "European",
    "honda": "Japanese",
    "lexus": "Japanese",
    "mazda": "Japanese",
    "mitsubishi": "Japanese",
    "nissan": "Japanese",
    "subaru": "Japanese",
    "suzuki": "Japanese",
    "toyota": "Japanese",
    "hyundai": "Korean",
    "kia": "Korean",
    "byd": "Chinese",
    "chery": "Chinese",
    "geely": "Chinese",
    "mg": "Chinese",
    "gwm": "Chinese",
    "ford": "American",
    "jeep": "American",
    "tesla": "American",
}

_MAKE_DISPLAY_NAMES = {
    "audi": "Audi",
    "bmw": "BMW",
    "ford": "Ford",
    "honda": "Honda",
    "hyundai": "Hyundai",
    "jeep": "Jeep",
    "kia": "Kia",
    "lexus": "Lexus",
    "mazda": "Mazda",
    "mercedes": "Mercedes-Benz",
    "mercedes-benz": "Mercedes-Benz",
    "mini": "MINI",
    "mitsubishi": "Mitsubishi",
    "nissan": "Nissan",
    "skoda": "Skoda",
    "subaru": "Subaru",
    "suzuki": "Suzuki",
    "tesla": "Tesla",
    "toyota": "Toyota",
    "volkswagen": "Volkswagen",
    "vw": "Volkswagen",
}

_COLUMN_ALIASES = {
    "make": ("make", "brand"),
    "model": ("model", "make and model"),
    "trim": ("trim", "sub model", "submodel"),
    "year": ("year", "vehicle year"),
    "body_style": ("body style", "body_style"),
    "fuel_type": ("fuel type", "fuel_type"),
    "transmission": ("transmission",),
    "engine_size": ("engine size", "engine_size"),
    "torque_nm": ("torque", "torque nm", "torque_nm"),
    "doors": ("doors", "number of doors"),
    "seats": ("seats", "number of seats"),
    "safety_stars": ("safety stars", "safety_stars"),
    "safety_rating": ("safety rating", "safety_rating"),
    "fuel_consumption_l_100km": (
        "fuel consumption (litres per 100km)",
        "fuel consumption",
        "fuel_consumption_l_100km",
    ),
}

_NUMERIC_FIELDS = (
    "year",
    "engine_displacement_cc",
    "engine_displacement_l",
    "engine_power_kw",
    "torque_nm",
    "doors",
    "seats",
    "safety_stars",
    "safety_rating",
    "fuel_consumption_l_100km",
)

_PRIMARY_RANKING_FIELDS = (
    "engine_type",
    "engine_power_kw",
    "transmission",
    "seats",
    "fuel_consumption_l_100km",
    "safety_rating",
)


def _normalize_key(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip().lower())


def _clean_text(value: Any) -> Optional[str]:
    text = re.sub(r"\s+", " ", str(value or "").strip())
    return text or None


def _source_value(row: Mapping[str, Any], aliases: Sequence[str]) -> Any:
    normalized = {_normalize_key(key): value for key, value in row.items()}
    for alias in aliases:
        value = normalized.get(_normalize_key(alias))
        if value not in (None, ""):
            return value
    return None


def _parse_number(value: Any) -> Optional[float]:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return float(value)
    match = re.search(r"-?\d+(?:[.,]\d+)?", str(value))
    return float(match.group(0).replace(",", ".")) if match else None


def _parse_int(value: Any) -> Optional[int]:
    number = _parse_number(value)
    return int(round(number)) if number is not None else None


def _normalize_fuel_type(value: Any) -> Optional[str]:
    text = _clean_text(value)
    if not text:
        return None
    lower = text.lower()
    if "plug" in lower and "hybrid" in lower:
        return "PHEV"
    if "hybrid" in lower:
        return "Hybrid"
    if "electric" in lower or lower in {"ev", "bev"}:
        return "EV"
    if "diesel" in lower:
        return "Diesel"
    if "petrol" in lower or "gasoline" in lower:
        return "Petrol"
    return text


def _engine_type(fuel_type: Optional[str], engine_size: Any) -> str:
    raw = _normalize_key(engine_size)
    if fuel_type == "EV" or any(token in raw for token in ("electric", "ev", "bev")):
        return "electric"
    if fuel_type == "PHEV":
        return "plug_in_hybrid"
    if fuel_type == "Hybrid":
        return "hybrid"
    if fuel_type == "Diesel":
        return "diesel"
    return "combustion"


def powertrain_category(
    fuel_type: Optional[str], engine_type: Optional[str] = None
) -> Optional[str]:
    """Classify a vehicle by whether external charging is required."""
    normalized_fuel = _normalize_fuel_type(fuel_type)
    if engine_type == "electric" or normalized_fuel == "EV":
        return "ev"
    if engine_type == "plug_in_hybrid" or normalized_fuel == "PHEV":
        return "plug_in_hybrid"
    if normalized_fuel in {"Petrol", "Diesel", "Hybrid"}:
        return "non_ev"
    return None


def normalize_powertrain_preference(value: Any) -> str:
    """Normalize the four UI choices to a stable filter value."""
    preference = _normalize_key(value)
    if preference in {"", "any", "i don't know", "i dont know", "unknown"}:
        return "any"
    if preference in {"ev", "electric"}:
        return "ev"
    if preference in {"plug in hybrid", "plug-in hybrid", "plug_in_hybrid", "phev"}:
        return "plug_in_hybrid"
    if preference in {
        "non ev",
        "non-ev",
        "non_ev",
        "petrol",
        "diesel",
        "hybrid",
        "mild hybrid",
    }:
        return "non_ev"
    return "any"


def parse_engine_spec(value: Any, fuel_type: Any = None) -> Dict[str, Optional[float]]:
    """Parse the government ``Engine size`` field into cc displacement and kW power.

    Combustion and hybrid rows are formatted like ``1998cc turbo 135kW``;
    electric rows contain motor power only, for example ``315kW``.
    """
    normalized_fuel = _normalize_fuel_type(fuel_type)
    text = _normalize_key(value)
    displacement_match = re.search(r"(\d+(?:[.,]\d+)?)\s*(?:cc|cm3|cm³)\b", text)
    power_match = re.search(r"(\d+(?:[.,]\d+)?)\s*kw\b", text)
    displacement_cc = (
        float(displacement_match.group(1).replace(",", "."))
        if displacement_match
        else None
    )
    power_kw = (
        float(power_match.group(1).replace(",", ".")) if power_match else None
    )
    if _engine_type(normalized_fuel, value) == "electric":
        displacement_cc = None
    if displacement_cc is not None and not 500 <= displacement_cc <= 10000:
        displacement_cc = None
    if power_kw is not None and not 20 <= power_kw <= 2000:
        power_kw = None
    return {
        "engine_displacement_cc": round(displacement_cc, 1)
        if displacement_cc is not None
        else None,
        "engine_power_kw": round(power_kw, 1) if power_kw is not None else None,
    }


def parse_engine_displacement_l(value: Any, fuel_type: Any = None) -> Optional[float]:
    """Compatibility helper returning litres from the parsed combustion displacement."""
    displacement_cc = parse_engine_spec(value, fuel_type)["engine_displacement_cc"]
    return round(displacement_cc / 1000, 3) if displacement_cc is not None else None


def _parse_torque_nm(value: Any) -> Optional[float]:
    number = _parse_number(value)
    if number is None:
        return None
    text = _normalize_key(value)
    if "lb-ft" in text or "lb ft" in text or "lbf" in text:
        number *= 1.35582
    return round(number, 1) if 20 <= number <= 2500 else None


def _parse_consumption_l_100km(value: Any) -> Optional[float]:
    number = _parse_number(value)
    if number is None:
        return None
    text = _normalize_key(value)
    if "mpg" in text:
        number = 282.481 / number
    return round(number, 2) if 0 < number <= 40 else None


def _normalize_transmission(value: Any) -> Optional[str]:
    text = _normalize_key(value)
    if not text:
        return None
    if "cvt" in text:
        return "CVT"
    if "manual" in text:
        return "Manual"
    if "dct" in text or "dsg" in text or "dual" in text:
        return "Dual-clutch"
    if "auto" in text or "automatic" in text:
        return "Automatic"
    if "single" in text and "speed" in text:
        return "Single-speed"
    return _clean_text(value)


def brand_region(make: Any) -> str:
    return BRAND_REGIONS.get(_normalize_key(make), "Other")


def _infer_make_and_model(make: Optional[str], model: Optional[str]) -> Tuple[Optional[str], Optional[str]]:
    if make:
        normalized_make = _normalize_key(make)
        return _MAKE_DISPLAY_NAMES.get(normalized_make, make), model
    if not model:
        return None, None
    lower_model = model.lower()
    for candidate in sorted(BRAND_REGIONS, key=len, reverse=True):
        if lower_model == candidate or lower_model.startswith(f"{candidate} "):
            remaining_model = model[len(candidate) :].strip() or None
            return _MAKE_DISPLAY_NAMES.get(candidate, candidate.title()), remaining_model
    return None, model


def validate_source_columns(columns: Iterable[Any]) -> Dict[str, List[str]]:
    """Return missing retained columns and unknown supplied columns."""
    supplied = {_normalize_key(column) for column in columns}
    required = {_normalize_key(column) for column in RETAIN_COLUMNS}
    return {
        "missing_required_columns": sorted(required - supplied),
        "unknown_columns": sorted(supplied - required),
    }


def normalize_car_record(row: Mapping[str, Any], default_make: Optional[str] = None) -> Dict[str, Any]:
    """Convert one raw spreadsheet/JSON row into a typed, provenance-carrying record."""
    raw = {key: value for key, value in row.items()}
    make = _clean_text(_source_value(raw, _COLUMN_ALIASES["make"]) or default_make)
    model = _clean_text(_source_value(raw, _COLUMN_ALIASES["model"]))
    make, model = _infer_make_and_model(make, model)
    if make and model and model.lower().startswith(f"{make.lower()} "):
        model = model[len(make) :].strip() or None

    fuel_type = _normalize_fuel_type(_source_value(raw, _COLUMN_ALIASES["fuel_type"]))
    raw_engine_size = _source_value(raw, _COLUMN_ALIASES["engine_size"])
    engine_type = _engine_type(fuel_type, raw_engine_size)
    normalized_powertrain_category = powertrain_category(fuel_type, engine_type)
    engine_spec = parse_engine_spec(raw_engine_size, fuel_type)
    safety_stars = _parse_number(_source_value(raw, _COLUMN_ALIASES["safety_stars"]))
    safety_rating = _parse_number(_source_value(raw, _COLUMN_ALIASES["safety_rating"]))
    if safety_rating is None:
        safety_rating = safety_stars

    record: Dict[str, Any] = {
        "make": make,
        "model": model,
        "trim": _clean_text(_source_value(raw, _COLUMN_ALIASES["trim"])),
        "year": _parse_int(_source_value(raw, _COLUMN_ALIASES["year"])),
        "body_style": _clean_text(_source_value(raw, _COLUMN_ALIASES["body_style"])),
        "fuel_type": fuel_type,
        "engine_type": engine_type,
        "powertrain_category": normalized_powertrain_category,
        "engine_displacement_cc": engine_spec["engine_displacement_cc"],
        "engine_displacement_l": parse_engine_displacement_l(raw_engine_size, fuel_type),
        "engine_power_kw": engine_spec["engine_power_kw"],
        "torque_nm": _parse_torque_nm(_source_value(raw, _COLUMN_ALIASES["torque_nm"])),
        "transmission": _normalize_transmission(_source_value(raw, _COLUMN_ALIASES["transmission"])),
        "doors": _parse_int(_source_value(raw, _COLUMN_ALIASES["doors"])),
        "seats": _parse_int(_source_value(raw, _COLUMN_ALIASES["seats"])),
        "fuel_consumption_l_100km": _parse_consumption_l_100km(
            _source_value(raw, _COLUMN_ALIASES["fuel_consumption_l_100km"])
        ),
        "safety_stars": round(safety_stars, 1) if safety_stars is not None else None,
        "safety_rating": round(safety_rating, 1) if safety_rating is not None else None,
        "brand_region": brand_region(make),
        "brand_region_source": "curated_make_map" if make else "missing_make",
        "raw_values": raw,
    }
    for field in _NUMERIC_FIELDS:
        if field in record:
            record[f"{field}_imputed"] = False
            record[f"{field}_imputation_level"] = None
    return record


def _cohort_key(record: Mapping[str, Any], level: str) -> Tuple[Any, ...]:
    if level == "model":
        return record.get("make"), record.get("model")
    if level == "generation":
        return record.get("make"), record.get("model")
    if level == "make_cohort":
        return record.get("make"), record.get("body_style"), record.get("engine_type")
    return record.get("brand_region"), record.get("body_style"), record.get("engine_type")


def _representative(values: Sequence[Any]) -> Any:
    if all(isinstance(value, (int, float)) and not isinstance(value, bool) for value in values):
        return round(float(median(values)), 2)
    return Counter(values).most_common(1)[0][0]


def impute_missing_values(records: Sequence[Mapping[str, Any]]) -> List[Dict[str, Any]]:
    """Impute only from increasingly broad cohorts and record every imputation."""
    normalized = [dict(record) for record in records]
    levels = ("model", "generation", "make_cohort", "region_cohort")
    fields = ("engine_displacement_cc", "engine_displacement_l", "engine_power_kw", "torque_nm", "transmission", "doors", "seats", "fuel_consumption_l_100km", "safety_stars", "safety_rating")
    profiles: Dict[str, Dict[str, Dict[Tuple[Any, ...], Any]]] = defaultdict(lambda: defaultdict(dict))

    for level in levels:
        buckets: Dict[str, Dict[Tuple[Any, ...], List[Any]]] = defaultdict(lambda: defaultdict(list))
        for record in normalized:
            key = _cohort_key(record, level)
            if any(value is None for value in key):
                continue
            for field in fields:
                value = record.get(field)
                if value is not None:
                    buckets[field][key].append(value)
        for field, field_buckets in buckets.items():
            profiles[level][field] = {
                key: _representative(values) for key, values in field_buckets.items()
            }

    for record in normalized:
        for field in fields:
            if record.get(field) is not None:
                continue
            if field in {"engine_displacement_cc", "engine_displacement_l"} and record.get("engine_type") == "electric":
                continue
            for level in levels:
                value = profiles[level][field].get(_cohort_key(record, level))
                if value is not None:
                    record[field] = value
                    record[f"{field}_imputed"] = True
                    record[f"{field}_imputation_level"] = level
                    break
    return normalized


def ranking_eligibility(record: Mapping[str, Any]) -> Dict[str, Any]:
    """Keep uncertain records visible to data QA but out of a misleading top ten."""
    missing = [field for field in _PRIMARY_RANKING_FIELDS if record.get(field) is None]
    imputed_primary = [
        field for field in _PRIMARY_RANKING_FIELDS if record.get(f"{field}_imputed")
    ]
    confidence = max(0.0, 1.0 - len(missing) * 0.2 - len(imputed_primary) * 0.05)
    return {
        "eligible": not missing and confidence >= 0.7,
        "missing_critical_fields": missing,
        "imputed_critical_fields": imputed_primary,
        "confidence": round(confidence, 2),
    }


def build_quality_report(records: Sequence[Mapping[str, Any]]) -> Dict[str, Any]:
    missingness = {
        field: sum(record.get(field) is None for record in records)
        for field in _NUMERIC_FIELDS
    }
    eligibility = [ranking_eligibility(record) for record in records]
    powertrain_counts = Counter(
        record.get("powertrain_category") or "unknown" for record in records
    )
    return {
        "record_count": len(records),
        "missing_value_counts": missingness,
        "powertrain_category_counts": dict(powertrain_counts),
        "eligible_record_count": sum(item["eligible"] for item in eligibility),
        "ineligible_reasons": Counter(
            reason for item in eligibility for reason in item["missing_critical_fields"]
        ),
    }


def preprocess_records(rows: Sequence[Mapping[str, Any]], default_make: Optional[str] = None) -> Tuple[List[Dict[str, Any]], Dict[str, Any]]:
    """Normalize, impute, and annotate a batch of source rows."""
    normalized = [normalize_car_record(row, default_make) for row in rows]
    enriched = impute_missing_values(normalized)
    for record in enriched:
        record["data_quality"] = ranking_eligibility(record)
    return enriched, build_quality_report(enriched)


def load_xlsx_rows(path: Path | str) -> List[Dict[str, Any]]:
    """Load the active worksheet into dictionaries without applying transformations."""
    from openpyxl import load_workbook

    workbook = load_workbook(Path(path), read_only=True, data_only=True)
    worksheet = workbook.active
    rows = worksheet.iter_rows(values_only=True)
    headers = next(rows, None)
    if not headers:
        return [] 
    normalized_headers = [str(header).strip() if header is not None else "" for header in headers]
    return [
        {
            header: value
            for header, value in zip(normalized_headers, values)
            if header
        }
        for values in rows
        if any(value not in (None, "") for value in values)
    ]


def preprocess_workbook(
    path: Path | str, default_make: Optional[str] = None
) -> Tuple[List[Dict[str, Any]], Dict[str, Any]]:
    """Load and preprocess one NZ vehicle source workbook."""
    workbook_path = Path(path)
    make = default_make or workbook_path.stem
    rows = load_xlsx_rows(workbook_path)
    records, report = preprocess_records(rows, default_make=make)
    report["source_path"] = str(workbook_path)
    report["source_column_validation"] = validate_source_columns(
        rows[0].keys() if rows else []
    )
    return records, report


def preprocess_workbooks(paths: Iterable[Path | str]) -> Tuple[List[Dict[str, Any]], Dict[str, Any]]:
    """Preprocess many workbooks and aggregate their data-quality reports."""
    all_records: List[Dict[str, Any]] = []
    per_workbook: Dict[str, Dict[str, Any]] = {}
    for path in sorted(Path(path) for path in paths):
        records, report = preprocess_workbook(path)
        all_records.extend(records)
        per_workbook[path.stem] = report

    aggregate = build_quality_report(all_records)
    aggregate["workbook_count"] = len(per_workbook)
    aggregate["per_workbook"] = per_workbook
    return all_records, aggregate


def write_processed_dataset(
    records: Sequence[Mapping[str, Any]],
    report: Mapping[str, Any],
    output_path: Path | str,
    report_path: Path | str,
) -> None:
    """Write normalized records and metadata as separate JSON artifacts."""
    output = Path(output_path)
    metadata = Path(report_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    metadata.parent.mkdir(parents=True, exist_ok=True)
    output.write_text(json.dumps(list(records), indent=2), encoding="utf-8")
    metadata.write_text(json.dumps(dict(report), indent=2), encoding="utf-8")
